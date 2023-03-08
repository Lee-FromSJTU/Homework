import openpyxl
import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt
from scipy import signal
from scipy.misc import central_diff_weights


def thermocline_detection(filename, filter_type='None'):
    """
    Parameters
    ----------
    filename: str
        The path to open or a file-like object.
    filter_type: str
        The type of filter. Default is 'None', which means no filtering.
    """
    wb = openpyxl.load_workbook(filename)
    ws = wb['Sheet1']

    depth = [ws.cell(row=i, column=1).value for i in range(2, 475)]
    temp = [ws.cell(row=i, column=2).value for i in range(2, 475)]

    if filter_type == 'lowpass':
        temp = lowpass_filtering(temp)

    grad = gradient(depth, temp)
    ag = sum(grad) / len(grad)

    flag = [abs(grad[i]) >= abs(ag) for i in range(len(grad))]

    [top, bottom] = detect(depth, flag)
    print(f'检测到{filename}中温跃层位置为{top}~{bottom}m')

    fig = plt.figure(figsize=(12, 8))
    ax = fig.add_subplot(1, 2, 1)
    td_data1 = np.array([[temp[i], depth[i]] for i in range(len(temp)) if bottom < depth[i] < top])
    td_data2 = np.array([[temp[i], depth[i]] for i in range(len(temp)) if not (bottom < depth[i] < top)])
    ax.plot(td_data1[:, 0], td_data1[:, 1], 'o', ms=5, mfc='w', mec='g', mew=1, alpha=0.85)
    ax.plot(td_data2[:, 0], td_data2[:, 1], 'o', ms=5, mfc='w', mec='k', mew=1, alpha=0.25)
    ax.plot([min(temp), max(temp)], [top, top], '--', color='r')
    ax.plot([min(temp), max(temp)], [bottom, bottom], '--', color='b')
    ax.grid()
    ax.set_xlabel(r'Temperature [$^\circ$C]')
    ax.set_ylabel('z [m]')
    ax.set_title('Temperature-Depth Profile')
    ax = fig.add_subplot(1, 2, 2)
    gd_data1 = np.array([[grad[i], depth[i]] for i in range(len(grad)) if flag[i]])
    gd_data2 = np.array([[grad[i], depth[i]] for i in range(len(grad)) if not flag[i]])
    ax.plot(gd_data1[:, 0], gd_data1[:, 1], 'o', ms=5, mfc='w', mec='r', mew=1, alpha=0.75)
    ax.plot(gd_data2[:, 0], gd_data2[:, 1], 'o', ms=5, mfc='w', mec='b', mew=1, alpha=0.75)
    ax.plot([ag, ag], [0, -40], '--', color='k', alpha=0.75)
    ax.grid()
    ax.set_xlabel(r'Temperature Gradient, -dT/dz[$^\circ$C/m]')
    ax.set_ylabel('z [m]')
    ax.set_title('Temperature Gradient-Depth Profile')
    fig.tight_layout()
    plt.show()

    return None


def gradient(x, y):
    """
    Parameters
    ----------
    x: list
        Independent variable
    y: list
        Dependent variable

    return
    ------
    g: list
        Forward difference
    """
    g = [(y[i + 1] - y[i]) / (x[i + 1] - x[i]) for i in range(len(x) - 1)]
    return g


def detect(depth, flag):
    """
    Parameters
    ----------
    depth: list
        Depth data
    flag: list
        Whether the temperature change rate exceeds the average value

    return
    ------
    boundary: list
        Top and bottom of thermocline.
    """
    t = []
    b = []
    if flag[0]:
        t.append(depth[0])
    for i in range(1, len(flag) - 1):
        if flag[i] and not flag[i - 1]:
            t.append(depth[i])
        if flag[i] and not flag[i + 1]:
            b.append(depth[i])
    if len(t) < len(b):
        t.append(depth[-1])
    elif len(t) > len(b):
        b.append(depth[-1])
    length = np.array([t[i] - b[i] for i in range(len(t))])
    idx = np.argmax(length)
    boundary = [t[idx], b[idx]]
    return boundary


def lowpass_filtering(temp):
    """
    Parameters
    ----------
    temp: list
        Temperature data

    return
    ------
    t: ndarray
        Temperature data after lowpass filtering
    """
    b, a = signal.butter(8, 0.05, 'lowpass')
    t = signal.filtfilt(b, a, temp)
    return t


if __name__ == '__main__':
    thermocline_detection('温深数据1.xlsx')
    thermocline_detection('温深数据2.xlsx', 'lowpass')
